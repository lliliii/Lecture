# -*- coding:utf-8 -*-

# builtin module
import os
import re
import sys

# install module
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def analyzer(combo_list:bytes)->list:
    email_regex = re.compile('([a-zA-Z0-9][a-zA-Z0-9._%+*-]+@[a-zA-Z0-9.-]+\\.[a-zA-Z\\d]{2,6})')
    mail_accounts = {}

    # 한줄 한줄 돌면서 이메일 카운팅
    for combo in combo_list:
        combo = combo.decode('utf-8', 'ignore').rstrip('\n')
        email = email_regex.search(combo)
        
        # 정규표현식에 이메일이 매칭될 경우
        if email:
            domain = email.group().split('@')[1]

            # 딕셔너리에 업데이트
            mail_accounts.setdefault( domain, 0)
            mail_accounts[domain] += 1

    # 갯수 순 정렬
    results = dict( sorted( mail_accounts.items(), key=lambda item: item[1], reverse=True ) )
    return results

def create_excel( filepath:str, combolist:list )->None:
    excel_workbook = Workbook()
    excel_worksheet = excel_workbook.active
    excel_worksheet.freeze_panes = 'A2'
    excel_worksheet.append(['Domain', 'Count', 'Korea'])

    excel_worksheet.cell(row=1, column=1).fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')
    excel_worksheet.cell(row=1, column=2).fill = PatternFill(start_color='ffff00', end_color='ffff00', fill_type='solid')

    # garbage 사용 안함, 
    for domain, count in combolist.items():
        is_kr = False
        if domain.endswith('.kr'):
            is_kr = True
        excel_worksheet.append([domain, count, is_kr])

    excel_worksheet.auto_filter.ref = 'A1:C1' # 엑셀 첫줄 필터
    excel_workbook.save( filename=f'{filepath}.xlsx' ) # 워크시트 저장

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print( 'usage: python3 combo.py {filepath}' )
        sys.exit()

    filepath = sys.argv[1]
    if os.path.isfile(filepath):
        with open(filepath, 'rb') as f:
            combolist = analyzer( f )
            create_excel( filepath, combolist )
